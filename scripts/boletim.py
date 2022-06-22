import pandas as pd             # manipulação dos dados tabulares e scraping
import numpy as np              # manipulação séries temporais
import ssl                      # conexão de segurança http
import datetime as dt           # manipulação de objetos com data
import urllib                   # pacote para requerimento de download
import matplotlib.pyplot as plt # pacote para criar gráficos
import openpyxl as oxl          # pacote para manipulação de arquivos .xlsx (preenchimento)

ssl._create_default_https_context = ssl._create_unverified_context

url_brent = 'http://www.ipeadata.gov.br/ExibeSerie.aspx?stub=1&serid1650971490=1650971490&PerID=1&PerMetodo1650971490_0=LST&module=M'
url_ipca = 'http://www.ipeadata.gov.br/ExibeSerie.aspx?serid=36482'
url_inpc = 'http://www.ipeadata.gov.br/ExibeSerie.aspx?serid=36472'
url_igpdi = 'http://www.ipeadata.gov.br/ExibeSerie.aspx?serid=33593'
url_igpm = 'http://www.ipeadata.gov.br/ExibeSerie.aspx?serid=37796'

pagina_brent = pd.read_html(url_brent)
pagina_ipca = pd.read_html(url_ipca)
pagina_inpc = pd.read_html(url_inpc)
pagina_igpdi = pd.read_html(url_igpdi)
pagina_igpm = pd.read_html(url_igpm)

url_precos_mensal = 'https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/precos/precos-revenda-e-de-distribuicao-combustiveis/shlp/mensal/mensal-municipios-desde-jan2013.xlsx'
download_precos_mensal = urllib.request.urlretrieve(url_precos_mensal, 'dados/precos_mensal.xlsx')

url_regioesbr_semanal = 'https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/precos/precos-revenda-e-de-distribuicao-combustiveis/shlp/semanal/semanal-regioes-desde-2013.xlsx'
download_regioesbr_semanal = urllib.request.urlretrieve(url_regioesbr_semanal, 'dados/precos_semanal_regioesbr.xlsx')

ipca_subitem_atual = pd.ExcelFile('dados/ipca_subitem_atual.xls')                       #Importa arquivo excel completo
ipca_subitem_mensal_atual = pd.read_excel(ipca_subitem_atual, 'MENSAL SUBITEM IPCA')    #Lê a planilha referente a variação mensal
ipca_subitem_acum_atual = pd.read_excel(ipca_subitem_atual, 'ACUM SUBITEM IPCA')        #Lê a planilha referente ao acumulado
ipca_subitem_acum12m_atual = pd.read_excel(ipca_subitem_atual, 'ACUM 12M SUBITEM IPCA') #Lê a planilha referente ao acum 12 meses
ipca_subitem_pesos_atual = pd.read_excel(ipca_subitem_atual, 'PESOS SUBITEM IPCA')      #Lê a planilha referente ao peso dos itens

ipca_subitem_anterior = pd.ExcelFile('dados/ipca_subitem_anterior.xls') 
ipca_subitem_mensal_anterior = pd.read_excel(ipca_subitem_anterior, 'MENSAL SUBITEM IPCA')    #Lê a planilha referente a variação mensal

inpc_subitem_atual = pd.ExcelFile('dados/inpc_subitem_atual.xls')                       #Importa arquivo excel completo
inpc_subitem_mensal_atual = pd.read_excel(inpc_subitem_atual, 'MENSAL SUBITEM INPC')    #Lê a planilha referente a variação mensal
inpc_subitem_acum_atual = pd.read_excel(inpc_subitem_atual, 'ACUM SUBITEM INPC')        #Lê a planilha referente ao acumulado
inpc_subitem_acum12m_atual = pd.read_excel(inpc_subitem_atual, 'ACUM 12M SUBITEM INPC') #Lê a planilha referente ao acum 12 meses
inpc_subitem_pesos_atual = pd.read_excel(inpc_subitem_atual, 'PESOS SUBITEM INPC')      #Lê a planilha referente ao peso dos itens

inpc_subitem_acum_anterior = pd.ExcelFile('dados/inpc_subitem_anterior.xls')                       #Importa arquivo excel completo
inpc_subitem_mensal_anterior = pd.read_excel(inpc_subitem_acum_anterior, 'MENSAL SUBITEM INPC')    #Lê a planilha referente a variação mensal

brent = pagina_brent[2]
brent.columns = ['Data','Preço - Brent (FOB)']
brent = brent.drop(0)
brent['Preço - Brent (FOB)'] = pd.to_numeric(brent['Preço - Brent (FOB)'])
brent['Preço - Brent (FOB)'] = brent['Preço - Brent (FOB)']/100
brent['Data'] = pd.to_datetime(brent['Data'])
brent.set_index('Data', inplace=True)

ipca = pagina_ipca[2]
ipca.columns = ['Data','IPCA Indice']
ipca = ipca.drop(range(0,105))
data = ipca['Data']
ipca['IPCA Indice'] = ipca['IPCA Indice'].str.replace('.','')
ipca['IPCA Indice'] = ipca['IPCA Indice'].str.replace(',','.')
ipca['IPCA Indice'] = pd.to_numeric(ipca['IPCA Indice'])
ipca['Data'] = pd.to_datetime(ipca['Data'])
ipca.set_index('Data', inplace=True)

ipca['Variação Mensal'] = 0.0
for i in range(-1, -21, -1):
    ipca['Variação Mensal'][i] = ((ipca['IPCA Indice'][i] / ipca['IPCA Indice'][i-1]) - 1)*100

ipca['Acumulado 12 meses'] = 0.0
for i in range(-1, -201, -1):
    ipca['Acumulado 12 meses'][i] = ((ipca['IPCA Indice'][i] / ipca['IPCA Indice'][i-12]) - 1)*100

ipca['Acumulado no Ano'] = 0.0
for i in range(401, len(ipca)):
    ipca['Acumulado no Ano'][i] = (ipca['IPCA Indice'][i] / ipca['IPCA Indice'][400] - 1)*100

ipca = ipca.round(2)

inpc = pagina_inpc[2]
inpc.columns = ['Data','INPC Indice']
inpc = inpc.drop(range(0,105))

inpc['INPC Indice'] = inpc['INPC Indice'].str.replace('.','')
inpc['INPC Indice'] = inpc['INPC Indice'].str.replace(',','.')
inpc['INPC Indice'] = pd.to_numeric(inpc['INPC Indice'])
inpc['Data'] = pd.to_datetime(inpc['Data'])
inpc.set_index('Data', inplace=True)

inpc['Variação Mensal'] = 0.0
for i in range(-1, -21, -1):
    inpc['Variação Mensal'][i] = ((inpc['INPC Indice'][i] / inpc['INPC Indice'][i-1]) - 1)*100

inpc['Acumulado 12 meses'] = 0.0
for i in range(-1, -201, -1):
    inpc['Acumulado 12 meses'][i] = ((inpc['INPC Indice'][i] / inpc['INPC Indice'][i-12]) - 1)*100

inpc['Acumulado no Ano'] = 0.0
for i in range(410, len(inpc)):
    inpc['Acumulado no Ano'][i] = (inpc['INPC Indice'][i] / inpc['INPC Indice'][409] - 1)*100

inpc = inpc.round(2)

#IPCA
ipca_subitem_mensal_atual.iloc[3][0]='ITENS'
ipca_subitem_acum_atual.iloc[3][0]='ITENS'
ipca_subitem_acum12m_atual.iloc[3][0]='ITENS'
ipca_subitem_pesos_atual.iloc[3][0]='ITENS'

ipca_subitem_mensal_atual.columns = ipca_subitem_mensal_atual.iloc[3]
ipca_subitem_acum_atual.columns = ipca_subitem_mensal_atual.iloc[3]
ipca_subitem_acum12m_atual.columns = ipca_subitem_mensal_atual.iloc[3]
ipca_subitem_pesos_atual.columns = ipca_subitem_mensal_atual.iloc[3]

ipca_subitem_mensal_atual.reset_index(inplace=True)
ipca_subitem_acum_atual.reset_index(inplace=True)
ipca_subitem_acum12m_atual.reset_index(inplace=True)
ipca_subitem_pesos_atual.reset_index(inplace=True)

ipca_subitem_mensal_atual.drop([0,1,2,3,4], inplace=True)
ipca_subitem_acum_atual.drop([0,1,2,3,4], inplace=True)
ipca_subitem_acum12m_atual.drop([0,1,2,3,4], inplace=True)
ipca_subitem_pesos_atual.drop([0,1,2,3,4], inplace=True)

ipca_subitem_mensal_anterior.iloc[3][0]='ITENS'
ipca_subitem_mensal_anterior.columns = ipca_subitem_mensal_anterior.iloc[3]
ipca_subitem_mensal_anterior.reset_index(inplace=True)
ipca_subitem_mensal_anterior.drop([0,1,2,3,4], inplace=True)

#INPC
inpc_subitem_mensal_atual.iloc[3][0]='ITENS'
inpc_subitem_acum_atual.iloc[3][0]='ITENS'
inpc_subitem_acum12m_atual.iloc[3][0]='ITENS'
inpc_subitem_pesos_atual.iloc[3][0]='ITENS'

inpc_subitem_mensal_atual.columns = inpc_subitem_mensal_atual.iloc[3]
inpc_subitem_acum_atual.columns = inpc_subitem_mensal_atual.iloc[3]
inpc_subitem_acum12m_atual.columns = inpc_subitem_mensal_atual.iloc[3]
inpc_subitem_pesos_atual.columns = inpc_subitem_mensal_atual.iloc[3]

inpc_subitem_mensal_atual.reset_index(inplace=True)
inpc_subitem_acum_atual.reset_index(inplace=True)
inpc_subitem_acum12m_atual.reset_index(inplace=True)
inpc_subitem_pesos_atual.reset_index(inplace=True)

inpc_subitem_mensal_atual.drop([0,1,2,3,4], inplace=True)
inpc_subitem_acum_atual.drop([0,1,2,3,4], inplace=True)
inpc_subitem_acum12m_atual.drop([0,1,2,3,4], inplace=True)
inpc_subitem_pesos_atual.drop([0,1,2,3,4], inplace=True)

inpc_subitem_mensal_anterior.iloc[3][0]='ITENS'
inpc_subitem_mensal_anterior.columns = inpc_subitem_mensal_anterior.iloc[3]
inpc_subitem_mensal_anterior.reset_index(inplace=True)
inpc_subitem_mensal_anterior.drop([0,1,2,3,4], inplace=True)

igpdi = pagina_igpdi[2]
igpdi.columns = ['Data','IGP-DI Indice']
igpdi = igpdi.drop(range(0,105))

igpdi['IGP-DI Indice'] = igpdi['IGP-DI Indice'].str.replace(',','.')
igpdi['IGP-DI Indice'] = igpdi['IGP-DI Indice'].str.replace('.','')
igpdi['IGP-DI Indice'] = pd.to_numeric(igpdi['IGP-DI Indice']) / 10000
igpdi['Data'] = pd.to_datetime(igpdi['Data'])
igpdi.set_index('Data', inplace=True)

igpdi['Variação Mensal'] = 0.0
for i in range(-1, -21, -1):
    igpdi['Variação Mensal'][i] = ((igpdi['IGP-DI Indice'][i] / igpdi['IGP-DI Indice'][i-1]) - 1)*100

igpdi['Acumulado 12 meses'] = 0.0
for i in range(-1, -100, -1):
    igpdi['Acumulado 12 meses'][i] = ((igpdi['IGP-DI Indice'][i] / igpdi['IGP-DI Indice'][i-12]) - 1)*100

igpdi['Acumulado no Ano'] = 0.0
for i in range(832, len(igpdi)):
    igpdi['Acumulado no Ano'][i] = (igpdi['IGP-DI Indice'][i] / igpdi['IGP-DI Indice'][831] - 1)*100

igpdi = igpdi.round(2)

igpm = pagina_igpm[2]
igpm.columns = ['Data','IGP-M Indice']
igpm = igpm.drop(range(0,3))

igpm['IGP-M Indice'] = igpm['IGP-M Indice'].str.replace(',','.')
igpm['IGP-M Indice'] = igpm['IGP-M Indice'].str.replace('.','')
igpm['IGP-M Indice'] = pd.to_numeric(igpm['IGP-M Indice']) / 10000
igpm['Data'] = pd.to_datetime(igpm['Data'])
igpm.set_index('Data', inplace=True)

igpm['Variação Mensal'] = 0.0
for i in range(-1, -21, -1):
    igpm['Variação Mensal'][i] = ((igpm['IGP-M Indice'][i] / igpm['IGP-M Indice'][i-1]) - 1)*100

igpm['Acumulado 12 meses'] = 0.0
for i in range(-1, -100, -1):
    igpm['Acumulado 12 meses'][i] = ((igpm['IGP-M Indice'][i] / igpm['IGP-M Indice'][i-12]) - 1)*100

igpm['Acumulado no Ano'] = 0.0
for i in range(390, len(igpm)):
    igpm['Acumulado no Ano'][i] = (igpm['IGP-M Indice'][i] / igpm['IGP-M Indice'][389] - 1)*100

igpm = igpm.round(2)

precos_semanal_regioesbr = pd.read_excel('dados/precos_semanal_regioesbr.xlsx', skiprows=17)

precos_mensal = pd.read_excel('dados/precos_mensal.xlsx', skiprows=16)

precos_semanal_regioesbr['DATA INICIAL'] = pd.to_datetime(precos_semanal_regioesbr['DATA INICIAL'])
precos_semanal_regioesbr['DATA FINAL'] = pd.to_datetime(precos_semanal_regioesbr['DATA FINAL'])
# precos_semanal_regioesbr['DATA INICIAL'] = pd.to_datetime(precos_semanal_regioesbr['DATA INICIAL'], unit='D', origin=pd.Timestamp('1900-01-01'))
# precos_semanal_regioesbr['DATA FINAL'] = pd.to_datetime(precos_semanal_regioesbr['DATA FINAL'], unit='D', origin=pd.Timestamp('1900-01-01'))

ipca.iloc[-12:,[1,2]]
plt.style.use('seaborn-whitegrid')

ax, fig = plt.subplots(figsize=(17, 9))
fig.plot(ipca.iloc[-12:,[1]], label = 'Mensal', linewidth=2, marker='o')
fig.plot(ipca.iloc[-12:,[2]], label = 'Acumulada em 12 Meses', linewidth=2, marker='o')

plt.grid(b=None)

plt.legend(
    title='Variação:',
    title_fontsize=14,
    loc='upper left',
    ncol = 1,
    bbox_to_anchor=(0.70,0.65),
    frameon=True,
    facecolor='white',
    shadow=True,
    fontsize=15)
fig.set_xlabel('Meses', fontsize=15)
fig.set_ylabel('Variação (%)', fontsize=15)
plt.yticks(fontsize=15)
str_month_list = list(ipca.index[-12:].strftime("%m/%Y"))
plt.xticks(ipca.iloc[-12:].index,str_month_list, fontsize=15)

# Valores em cada ponto do gráfico
for x,y in zip(ipca['Variação Mensal'][-12:].index,ipca['Variação Mensal'][-12:].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                 (x,y), # this is the point to label
                 textcoords="offset points", # how to position the text
                 xytext=(0,10), # distance from text to points (x,y)
                 ha='center',
                 size=15) # horizontal alignment can be left, right or center

for x,y in zip(ipca['Acumulado 12 meses'][-12:].index,ipca['Acumulado 12 meses'][-12:].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                 (x,y), # this is the point to label
                 textcoords="offset points", # how to position the text
                 xytext=(0,13), # distance from text to points (x,y)
                 ha='center',
                 size=15) # horizontal alignment can be left, right or center

plt.savefig('graficos/GRÁFICO 1 - IPCA 12M E MENSAL.svg', dpi=600)

aux_data = pd.DataFrame({'data':pd.date_range('2022-01-01', periods=12, freq='MS'), 'Values': range(0, 12)}) #variável auxiliar para criar meses futuros no gráfico da variação acumulada.

aux_data.set_index('data', inplace=True)

ipca.iloc[-12:,[3]]

plt.style.use('seaborn-whitegrid')
ax, fig = plt.subplots(figsize=(17, 9))
fig.plot(ipca.iloc[401:,[3]], marker='o', label = 'Acumulado no Ano', linewidth=2)
fig.plot(aux_data, alpha=0)
#fig.set_title('Variação Acumulado no Ano do IPCA')

plt.grid(False)

plt.legend(
    title='Variação:',
    title_fontsize=14,
    loc='upper left',
    ncol = 1,
    bbox_to_anchor=(0.05,0.97),
    frameon=True,
    facecolor='white',
    shadow=True,
    fontsize=15)
fig.set_xlabel('Meses', fontsize=15)
fig.set_ylabel('Variação (%)', fontsize=15)
plt.yticks(fontsize=12)
plt.xticks(fontsize=12)

str_month_list_acum = list(aux_data.index[-12:].strftime("%m/%Y"))
plt.xticks(aux_data.index,str_month_list_acum, fontsize=12)

# Valores em cada ponto do gráfico
for x,y in zip(ipca['Acumulado no Ano'][-12:].index,ipca['Acumulado no Ano'][-12:].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                 (x,y), # this is the point to label
                 textcoords="offset points", # how to position the text
                 xytext=(0,13), # distance from text to points (x,y)
                 ha='center',
                 size=15) # horizontal alignment can be left, right or center
plt.savefig('graficos/GRÁFICO 2 - IPCA NO ANO.svg', dpi=600)

regioes_dict = {
"RJ":'Rio de Janeiro',	
"POA":'Porto Alegre',
"BH":"Belo Horizonte",
"REC":"Recife",
"SP":"São Paulo",
"DF":"Brasília",
"BEL":"Belém",
"FOR":"Forteleza",
"SAL":"Salvador",
"CUR":"Curitiba",
"GOI":"Goiânia",
"VIT":"Vitória",
"CG":"Campo Grande",
"RB":"Rio Branco",
"SL":"São Luís",
"AJU":"Aracaju",
"NACIONAL":"Brasil"
}

pesos = [['RJ','POA','BH','REC','SP','DF','BEL','FOR','SAL','CUR','GOI','VIT','CG','RB','SL','AJU','NACIONAL'],
[9.42, 8.61, 9.69, 3.92, 32.28, 4.06, 3.94, 3.23, 5.99, 8.09, 4.17, 1.86, 1.57, 0.51, 1.62, 1.03, 100]]
regioes_ipca_peso_atual = pd.DataFrame(pesos)
regioes_ipca_peso_atual = regioes_ipca_peso_atual.transpose()
regioes_ipca_peso_atual.columns = ['Regiao','Peso']

regioes_ipca_var_anterior = ipca_subitem_mensal_anterior[ipca_subitem_mensal_anterior.ITENS == ' ÍNDICE GERAL']

regioes_ipca_var_atual = ipca_subitem_mensal_atual[ipca_subitem_mensal_atual.ITENS == ' ÍNDICE GERAL']

regioes_ipca_acum_atual = ipca_subitem_acum_atual[ipca_subitem_acum_atual.ITENS == ' ÍNDICE GERAL']

regioes_ipca_var_anterior = regioes_ipca_var_anterior.transpose().reset_index()
regioes_ipca_var_anterior.rename(columns={3:'Regiao',5:'Var ant'}, inplace=True)
regioes_ipca_var_anterior.drop(range(0,2), inplace=True)

regioes_ipca_var_atual = regioes_ipca_var_atual.transpose().reset_index()
regioes_ipca_var_atual.rename(columns={3:'Regiao',5:'Var atual'}, inplace=True)
regioes_ipca_var_atual.drop(range(0,2), inplace=True)

regioes_ipca_acum_atual = regioes_ipca_acum_atual.transpose().reset_index()
regioes_ipca_acum_atual.rename(columns={3:'Regiao',5:'Var acum'}, inplace=True)
regioes_ipca_acum_atual.drop(range(0,2), inplace=True)

regioes_ipca_12m_atual = ipca_subitem_acum12m_atual[ipca_subitem_acum12m_atual.ITENS == ' ÍNDICE GERAL']

regioes_ipca_12m_atual = regioes_ipca_12m_atual.transpose().reset_index()
regioes_ipca_12m_atual.rename(columns={3:'Regiao',5:'Var 12m'}, inplace=True)
regioes_ipca_12m_atual.drop(range(0,2), inplace=True)

subitens_regioes_ipca = pd.merge(regioes_ipca_var_anterior, regioes_ipca_var_atual, how='outer')

subitens_regioes_ipca = pd.merge(subitens_regioes_ipca, regioes_ipca_acum_atual, how='outer')

subitens_regioes_ipca = pd.merge(subitens_regioes_ipca, regioes_ipca_12m_atual, how='outer')

subitens_regioes_ipca = pd.merge(subitens_regioes_ipca, regioes_ipca_peso_atual, how='outer')

subitens_regioes_ipca['Regiao'] = [regioes_dict[resp] for resp in subitens_regioes_ipca.Regiao]

linha_brasil = subitens_regioes_ipca[subitens_regioes_ipca['Regiao'] == 'Brasil']

subitens_regioes_ipca.drop(16, inplace=True)

subitens_regioes_ipca.sort_values('Var atual', ascending=False, inplace=True)

subitens_regioes_ipca = pd.concat([subitens_regioes_ipca, linha_brasil])

workbook = oxl.load_workbook(filename='tabela_regioes.xlsx')

sheet = workbook.active

#Preenche a tabela das categorias do ipca de curitiba
for i in range(0, 17):
    sheet[f"B{i+4}"] = subitens_regioes_ipca.iloc[i,5]
    sheet[f"A{i+4}"] = subitens_regioes_ipca.iloc[i,0]
    sheet[f"C{i+4}"] = subitens_regioes_ipca.iloc[i,1]
    sheet[f"D{i+4}"] = subitens_regioes_ipca.iloc[i,2]
    sheet[f"E{i+4}"] = subitens_regioes_ipca.iloc[i,3]
    sheet[f"F{i+4}"] = subitens_regioes_ipca.iloc[i,4]

workbook.save(filename="tabelas\IPCA REGIÕES.xlsx")

ipca_cur_mensal_anterior = ipca_subitem_mensal_anterior[ipca_subitem_mensal_anterior.index.isin([6, 194, 234, 275, 317, 350, 396, 426, 451])]
ipca_cur_mensal_atual = ipca_subitem_mensal_atual[ipca_subitem_mensal_atual.index.isin([6, 194, 234, 275, 317, 350, 396, 426, 451])]
ipca_cur_acum_atual = ipca_subitem_acum_atual[ipca_subitem_acum_atual.index.isin([6, 194, 234, 275, 317, 350, 396, 426, 451])]
ipca_cur_acum12m_atual = ipca_subitem_acum12m_atual[ipca_subitem_acum12m_atual.index.isin([6, 194, 234, 275, 317, 350, 396, 426, 451])]

ipca_cur_mensal_anterior.rename(columns={'CUR':'CUR VAR ANT'}, inplace=True)
ipca_cur_mensal_atual.rename(columns={'CUR':'CUR VAR'}, inplace=True)
ipca_cur_acum_atual.rename(columns={'CUR':'CUR ACUM'}, inplace=True)
ipca_cur_acum12m_atual.rename(columns={'CUR':'CUR 12M'}, inplace=True)

subitens_cur_ipca = pd.merge(ipca_cur_mensal_anterior[['ITENS', 'CUR VAR ANT']], ipca_cur_mensal_atual[['ITENS', 'CUR VAR']], how='outer')
subitens_cur_ipca = pd.merge(subitens_cur_ipca, ipca_cur_acum12m_atual[['ITENS', 'CUR 12M']], how='outer')
subitens_cur_ipca = pd.merge(subitens_cur_ipca, ipca_cur_acum_atual[['ITENS', 'CUR ACUM']], how='outer')
subitens_cur_ipca = subitens_cur_ipca.sort_values('CUR VAR', ascending=False).reset_index()

workbook = oxl.load_workbook(filename='tabela_curitiba.xlsx') #Substituir Valores no Excel

sheet = workbook.active

#Preenche a tabela das categorias do ipca de curitiba
for i in range(0, 9):
    sheet[f"A{i+4}"] = subitens_cur_ipca.iloc[i,1]
    sheet[f"B{i+4}"] = subitens_cur_ipca.iloc[i,2]
    sheet[f"C{i+4}"] = subitens_cur_ipca.iloc[i,3]
    sheet[f"D{i+4}"] = subitens_cur_ipca.iloc[i,4]
    sheet[f"E{i+4}"] = subitens_cur_ipca.iloc[i,5]

workbook.save(filename="tabelas\IPCA CURITIBA.xlsx")

inpc.iloc[-12:,[1,2]]
plt.style.use('seaborn-whitegrid')

ax, fig = plt.subplots(figsize=(13, 6))
fig.plot(inpc.iloc[-12:,[1]], label = 'Mensal', linewidth=2, marker='o')
fig.plot(inpc.iloc[-12:,[2]], label = 'Acumulada em 12 Meses', linewidth=2, marker='o')
#fig.set_title('Variação Mensal e Acumulada dos Últimos 12 meses do INPC')

plt.grid(False)
plt.legend(
    title='Variação:',
    title_fontsize=14,
    loc='upper left',
    ncol = 1,
    bbox_to_anchor=(0.65,0.60),
    frameon=True,
    facecolor='white',
    shadow=True,
    fontsize=15)
fig.set_xlabel('Meses', fontsize=15)
fig.set_ylabel('Variação (%)', fontsize=15)
plt.yticks(fontsize=12)
plt.xticks(fontsize=12)

str_month_list = list(inpc.index[-12:].strftime("%m/%Y"))
plt.xticks(inpc.index[-12:],str_month_list, fontsize=12)



# Valores em cada ponto do gráfico
for x,y in zip(inpc['Variação Mensal'][-12:].index,inpc['Variação Mensal'][-12:].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                 (x,y), # this is the point to label
                 textcoords="offset points", # how to position the text
                 xytext=(0,10), # distance from text to points (x,y)
                 ha='center',
                 size=15) # horizontal alignment can be left, right or center

for x,y in zip(inpc['Acumulado 12 meses'][-12:].index,inpc['Acumulado 12 meses'][-12:].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                 (x,y), # this is the point to label
                 textcoords="offset points", # how to position the text
                 xytext=(0,13), # distance from text to points (x,y)
                 ha='center',
                 size=15) # horizontal alignment can be left, right or center

plt.savefig('graficos/GRÁFICO 3 - INPC 12M E MENSAL.svg', dpi=600)

inpc.iloc[-12:,[3]]

plt.style.use('seaborn-whitegrid')
ax, fig = plt.subplots(figsize=(13, 6))
fig.plot(inpc.iloc[410:,[3]], marker='o', label = 'Acumulado no Ano', linewidth=2)
fig.plot(aux_data, alpha=0)
#fig.set_title('Variação Acumulada no Ano INPC')

plt.grid(False)

plt.legend(
    title='Variação:',
    title_fontsize=14,
    loc='upper left',
    ncol = 1,
    bbox_to_anchor=(0.05,0.97),
    frameon=True,
    facecolor='white',
    shadow=True,
    fontsize=15)
fig.set_xlabel('Meses', fontsize=15)
fig.set_ylabel('Variação (%)', fontsize=15)
plt.yticks(fontsize=12)
plt.xticks(fontsize=12)

str_month_list_acum = list(aux_data.index[-12:].strftime("%m/%Y"))
plt.xticks(aux_data.index,str_month_list_acum, fontsize=12)

# Valores em cada ponto do gráfico
for x,y in zip(inpc['Acumulado no Ano'][-12:].index,inpc['Acumulado no Ano'][-12:].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                 (x,y), # this is the point to label
                 textcoords="offset points", # how to position the text
                 xytext=(0,13), # distance from text to points (x,y)
                 ha='center',
                 size=15) # horizontal alignment can be left, right or center

plt.savefig('graficos/GRÁFICO 4 - INPC NO ANO.svg')

regioes_inpc_var_anterior = inpc_subitem_mensal_anterior[inpc_subitem_mensal_anterior.ITENS == ' ÍNDICE GERAL']
regioes_inpc_var_atual = inpc_subitem_mensal_atual[inpc_subitem_mensal_atual.ITENS == ' ÍNDICE GERAL']
regioes_inpc_acum_atual = inpc_subitem_acum_atual[inpc_subitem_acum_atual.ITENS == ' ÍNDICE GERAL']

pesos = [['RJ','POA','BH','REC','SP','DF','BEL','FOR','SAL','CUR','GOI','VIT','CG','RB','SL','AJU','NACIONAL'],
[9.30,7.15,10.35,5.60,24.60,1.97,6.95,5.16,7.92,7.37,4.43,1.91,1.73,0.72,3.47,1.29,100]]

regioes_inpc_peso_atual = pd.DataFrame(pesos)
regioes_inpc_peso_atual = regioes_inpc_peso_atual.transpose()
regioes_inpc_peso_atual.columns = ['Regiao','Peso']

regioes_inpc_peso_atual = regioes_inpc_peso_atual.transpose().reset_index()
regioes_inpc_peso_atual.rename(columns={3:'Regiao',5:'Var'}, inplace=True)
regioes_inpc_peso_atual.drop(range(0,2), inplace=True)

regioes_inpc_var_anterior = regioes_inpc_var_anterior.transpose().reset_index()
regioes_inpc_var_anterior.rename(columns={3:'Regiao',5:'Var ant'}, inplace=True)
regioes_inpc_var_anterior.drop(range(0,2), inplace=True)

regioes_inpc_var_atual = regioes_inpc_var_atual.transpose().reset_index()
regioes_inpc_var_atual.rename(columns={3:'Regiao',5:'Var atual'}, inplace=True)
regioes_inpc_var_atual.drop(range(0,2), inplace=True)

regioes_inpc_acum_atual = regioes_inpc_acum_atual.transpose().reset_index()
regioes_inpc_acum_atual.rename(columns={3:'Regiao',5:'Var acum'}, inplace=True)
regioes_inpc_acum_atual.drop(range(0,2), inplace=True)

regioes_inpc_12m_atual = inpc_subitem_acum12m_atual[inpc_subitem_acum12m_atual.ITENS == ' ÍNDICE GERAL']

regioes_inpc_12m_atual = regioes_inpc_12m_atual.transpose().reset_index()
regioes_inpc_12m_atual.rename(columns={3:'Regiao',5:'Var 12m'}, inplace=True)
regioes_inpc_12m_atual.drop(range(0,2), inplace=True)

subitens_regioes_inpc = pd.merge(regioes_inpc_var_anterior, regioes_inpc_var_atual, how='outer') #Concatena as colunas com as variações

pesos = [['RJ','POA','BH','REC','SP','DF','BEL','FOR','SAL','CUR','GOI','VIT','CG','RB','SL','AJU','NACIONAL'],
[9.30,7.15,10.35,5.60,24.60,1.97,6.95,5.16,7.92,7.37,4.43,1.91,1.73,0.72,3.47,1.29,100]]
regioes_inpc_peso_atual = pd.DataFrame(pesos)
regioes_inpc_peso_atual = regioes_inpc_peso_atual.transpose()
regioes_inpc_peso_atual.columns = ['Regiao','Peso']

subitens_regioes_inpc = pd.merge(subitens_regioes_inpc, regioes_inpc_acum_atual, how='outer')
subitens_regioes_inpc = pd.merge(subitens_regioes_inpc, regioes_inpc_12m_atual, how='outer')
subitens_regioes_inpc = pd.merge(subitens_regioes_inpc, regioes_inpc_peso_atual, how='outer')
subitens_regioes_inpc['Regiao'] = [regioes_dict[resp] for resp in subitens_regioes_inpc.Regiao]

linha_brasil = subitens_regioes_inpc[subitens_regioes_inpc['Regiao'] == 'Brasil']

subitens_regioes_inpc.drop(16, inplace=True)

subitens_regioes_inpc.sort_values('Var atual', ascending=False, inplace=True)

subitens_regioes_inpc = pd.concat([subitens_regioes_inpc, linha_brasil])

workbook = oxl.load_workbook(filename='tabela_regioes.xlsx')

sheet = workbook.active

#Preenche a tabela das categorias do inpc de curitiba
for i in range(0, 17):
    sheet[f"B{i+4}"] = subitens_regioes_inpc.iloc[i,5]
    sheet[f"A{i+4}"] = subitens_regioes_inpc.iloc[i,0]
    sheet[f"C{i+4}"] = subitens_regioes_inpc.iloc[i,1]
    sheet[f"D{i+4}"] = subitens_regioes_inpc.iloc[i,2]
    sheet[f"E{i+4}"] = subitens_regioes_inpc.iloc[i,3]
    sheet[f"F{i+4}"] = subitens_regioes_inpc.iloc[i,4]

workbook.save(filename="tabelas/INPC REGIÕES.xlsx")

inpc_cur_mensal_anterior = inpc_subitem_mensal_anterior[inpc_subitem_mensal_anterior.index.isin([6, 184, 224, 264, 307, 341, 386, 416, 441])]
inpc_cur_mensal_atual = inpc_subitem_mensal_atual[inpc_subitem_mensal_atual.index.isin([6, 184, 224, 264, 307, 341, 386, 416, 441])]
inpc_cur_acum_atual = inpc_subitem_acum_atual[inpc_subitem_acum_atual.index.isin([6, 184, 224, 264, 307, 341, 386, 416, 441])]
inpc_cur_acum12m_atual = inpc_subitem_acum12m_atual[inpc_subitem_acum12m_atual.index.isin([6, 184, 224, 264, 307, 341, 386, 416, 441])]

#renomear colunas
inpc_cur_mensal_anterior.rename(columns={'CUR':'CUR VAR ANT'}, inplace=True)
inpc_cur_mensal_atual.rename(columns={'CUR':'CUR VAR'}, inplace=True)
inpc_cur_acum_atual.rename(columns={'CUR':'CUR ACUM'}, inplace=True)
inpc_cur_acum12m_atual.rename(columns={'CUR':'CUR 12M'}, inplace=True)

subitens_cur_inpc = pd.merge(inpc_cur_mensal_anterior[['ITENS', 'CUR VAR ANT']], inpc_cur_mensal_atual[['ITENS', 'CUR VAR']], how='outer') #concatenação das colunas para uma unica tabela
subitens_cur_inpc = pd.merge(subitens_cur_inpc, inpc_cur_acum12m_atual[['ITENS', 'CUR 12M']], how='outer')
subitens_cur_inpc = pd.merge(subitens_cur_inpc, inpc_cur_acum_atual[['ITENS', 'CUR ACUM']], how='outer')
subitens_cur_inpc = subitens_cur_inpc.sort_values('CUR VAR', ascending=False).reset_index()

subitens_cur_inpc['ITENS'] = subitens_cur_inpc['ITENS'].astype(str)

workbook = oxl.load_workbook(filename='tabela_curitiba.xlsx')

sheet = workbook.active

#Preenche a tabela das categorias do ipca de curitiba
for i in range(0, 9):
    sheet[f"A{i+4}"] = subitens_cur_inpc.iloc[i,1].strip().capitalize()
    sheet[f"B{i+4}"] = subitens_cur_inpc.iloc[i,2]
    sheet[f"C{i+4}"] = subitens_cur_inpc.iloc[i,3]
    sheet[f"D{i+4}"] = subitens_cur_inpc.iloc[i,5]
    sheet[f"E{i+4}"] = subitens_cur_inpc.iloc[i,4]

workbook.save(filename="tabelas/INPC CURITIBA.xlsx")

ax, fig = plt.subplots(figsize=(17, 9))

fig.plot(ipca.iloc[-12:,[1]], label = 'Mensal (IPCA)', linewidth=2, color='blue', marker='o')
fig.plot(ipca.iloc[-12:,[2]], label = 'Acumulada em 12 Meses (IPCA)', linewidth=2, color='blue', marker='o')

fig.plot(inpc.iloc[-12:,[1]], label = 'Mensal (INPC)', linewidth=2, color='green', marker='o')
fig.plot(inpc.iloc[-12:,[2]], label = 'Acumulada em 12 Meses (INPC)', linewidth=2, color='green', marker='o')

plt.grid(False)

plt.legend(
    title='Variação:',
    title_fontsize=14,
    loc='upper left',
    ncol = 1,
    bbox_to_anchor=(0.65,0.70),
    frameon=True,
    facecolor='white',
    shadow=True,
    fontsize=15)
fig.set_xlabel('Meses', fontsize=15)
fig.set_ylabel('Variação (%)', fontsize=15)
plt.yticks(fontsize=15)
plt.xticks(fontsize=15)

str_month_list = list(ipca.index[-12:].strftime("%m/%Y"))
plt.xticks(ipca.iloc[-12:].index,str_month_list, fontsize=15)

# Valores em cada ponto do gráfico
for x,y in zip(ipca['Variação Mensal'][-1:].index,ipca['Variação Mensal'][-1:].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                 (x,y), # this is the point to label
                 textcoords="offset points", # how to position the text
                 xytext=(0,-17), # distance from text to points (x,y)
                 ha='center',
                 size=15,
                 color='blue') # horizontal alignment can be left, right or center

for x,y in zip(ipca['Acumulado 12 meses'][-1:].index,ipca['Acumulado 12 meses'][-1:].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                 (x,y), # this is the point to label
                 textcoords="offset points", # how to position the text
                 xytext=(7,-15), # distance from text to points (x,y)
                 ha='center',
                 size=15,
                 color='blue') # horizontal alignment can be left, right or center

# Valores em cada ponto do gráfico
for x,y in zip(inpc['Variação Mensal'][-1:].index,inpc['Variação Mensal'][-1:].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                (x,y), # this is the point to label
                textcoords="offset points", # how to position the text
                xytext=(0,5), # distance from text to points (x,y)
                ha='center',
                size=15,
                color='green') # horizontal alignment can be left, right or center

for x,y in zip(inpc['Acumulado 12 meses'][-1:].index,inpc['Acumulado 12 meses'][-1:].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                (x,y), # this is the point to label
                textcoords="offset points", # how to position the text
                xytext=(0,5), # distance from text to points (x,y)
                ha='center',
                size=15,
                color='green') # horizontal alignment can be left, right or center

plt.savefig('graficos/GRÁFICO 5 - IPCA X INPC VAR 12M.svg', dpi=600)

precos_monitorados = [' ENERGIA ELÉTRICA RESIDENCIAL', ' ÓLEO DIESEL', ' PEDÁGIO', ' GÁS ENCANADO', ' TAXA DE ÁGUA E ESGOTO']
precos_monitorados_IPCA_cod = [199, 231, 232, 340, 348]
precos_monitorados_INPC_cod = [189, 221, 222, 330, 339]

inpc_subitem_acum12m_atual[inpc_subitem_acum12m_atual['index'].isin(precos_monitorados_INPC_cod)].replace('       -','0', inplace=True)

inpc_subitem_acum12m_atual[inpc_subitem_acum12m_atual['index'].isin(precos_monitorados_INPC_cod)].groupby(['ITENS','CUR']).sum().reset_index()

precos_monitorados_ipca = ipca_subitem_acum12m_atual.filter(['index','ITENS','CUR'])
precos_monitorados_ipca = ipca_subitem_acum12m_atual[ipca_subitem_acum12m_atual['index'].isin(precos_monitorados_IPCA_cod)].groupby(['ITENS','CUR']).sum().reset_index()

precos_monitorados_inpc = inpc_subitem_acum12m_atual.filter(['index','ITENS','CUR'])
precos_monitorados_inpc = inpc_subitem_acum12m_atual[inpc_subitem_acum12m_atual['index'].isin(precos_monitorados_INPC_cod)].groupby(['ITENS','CUR']).sum().reset_index()

X = precos_monitorados_ipca['ITENS'].values
ipca_monitorado = precos_monitorados_ipca['CUR'].values
inpc_monitorado = precos_monitorados_inpc['CUR'].values

X_axis = np.arange(len(X))

inpc_monitorado[1] = 0

ax, fig = plt.subplots(figsize=(17, 9))
baripca = plt.barh(X_axis - 0.2, ipca_monitorado, 0.4, label = 'IPCA', color='blue')
barinpc = plt.barh(X_axis + 0.2, inpc_monitorado, 0.4, label = 'INPC', color='green')
plt.yticks(X_axis, X)

plt.yticks(fontsize=15)
plt.xticks(fontsize=14)

plt.bar_label(baripca, fontsize=25)
plt.bar_label(barinpc, fontsize=25)

plt.grid(False)

plt.legend(
    title='Variação:',
    title_fontsize=20,
    loc='upper left',
    ncol = 1,
    bbox_to_anchor=(0.70,0.70),
    frameon=True,
    facecolor='white',
    shadow=True,
    fontsize=20)
fig.set_xlabel('Variação (%)', fontsize=20)

plt.yticks(fontsize=20)
plt.xticks(fontsize=20)

plt.tight_layout(pad=1)

plt.savefig('graficos/GRÁFICO 6 - PREÇOS MONITORADOS.svg', dpi=600)

ax, fig = plt.subplots(figsize=(17, 9))
fig.plot(igpdi.iloc[-12:,1], label = 'Mensal', linewidth=2, marker='o', color='blue')
fig.plot(igpdi.iloc[-12:,2], label = 'Acumulada em 12 Meses', linewidth=2, marker='o', color='green')
#fig.set_title('Variação Mensal e Acumulada dos Últimos 12 meses do IGP-DI')

plt.grid(False)

plt.legend(
    title='Variação:',
    title_fontsize=14,
    loc='upper left',
    ncol = 1,
    bbox_to_anchor=(0.70,0.97),
    frameon=True,
    facecolor='white',
    shadow=True,
    fontsize=15)
fig.set_xlabel('Meses', fontsize=20)
fig.set_ylabel('Variação (%)', fontsize=20)
plt.yticks(fontsize=15)
plt.xticks(fontsize=15)

str_month_list = list(igpdi.index[-12:].strftime("%m/%Y"))
plt.xticks(igpdi.iloc[-12:].index,str_month_list, fontsize=15)

# Valores em cada ponto do gráfico
for x,y in zip(igpdi['Variação Mensal'][-12:].index,igpdi['Variação Mensal'][-12:].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                (x,y), # this is the point to label
                textcoords="offset points", # how to position the text
                xytext=(0,10), # distance from text to points (x,y)
                ha='center',
                size=15) # horizontal alignment can be left, right or center

for x,y in zip(igpdi['Acumulado 12 meses'][-12:].index,igpdi['Acumulado 12 meses'][-12:].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                (x,y), # this is the point to label
                textcoords="offset points", # how to position the text
                xytext=(0,13), # distance from text to points (x,y)
                ha='center',
                size=15) # horizontal alignment can be left, right or center
plt.savefig('graficos/GRÁFICO 7 - IGP-DI.svg', dpi=600)

plt.style.use('seaborn-whitegrid')
ax, fig = plt.subplots(figsize=(17, 9))
fig.plot(igpm.iloc[-12:,1], label = 'Mensal', linewidth=2, marker='o', color='blue')
fig.plot(igpm.iloc[-12:,2], label = 'Acumulada em 12 Meses', linewidth=2, marker='o', color='green')
#fig.set_title('Variação Mensal e Acumulada dos Últimos 12 meses do IGP-M')

plt.grid(False)

plt.legend(
    title='Variação:',
    title_fontsize=14,
    loc='upper left',
    ncol = 1,
    bbox_to_anchor=(0.70,0.97),
    frameon=True,
    facecolor='white',
    shadow=True,
    fontsize=15)
fig.set_xlabel('Meses', fontsize=15)
fig.set_ylabel('Variação (%)', fontsize=15)
plt.yticks(fontsize=15)
plt.xticks(fontsize=15)

str_month_list = list(igpm.index[-12:].strftime("%m/%Y"))
plt.xticks(igpm.iloc[-12:].index,str_month_list, fontsize=15)

# Valores em cada ponto do gráfico
for x,y in zip(igpm['Variação Mensal'][-12:].index, igpm['Variação Mensal'][-12:].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                (x,y), # this is the point to label
                textcoords="offset points", # how to position the text
                xytext=(0,10), # distance from text to points (x,y)
                ha='center',
                size=15) # horizontal alignment can be left, right or center

for x,y in zip(igpm['Acumulado 12 meses'][-12:].index, igpm['Acumulado 12 meses'][-12:].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                (x,y), # this is the point to label
                textcoords="offset points", # how to position the text
                xytext=(0,13), # distance from text to points (x,y)
                ha='center',
                size=15) # horizontal alignment can be left, right or center

plt.savefig('graficos/GRÁFICO 8 - IGP-M.svg', dpi=600)

brent.dropna(inplace=True)

plt.style.use('seaborn-whitegrid')
ax, fig = plt.subplots(figsize=(17, 9))
fig.plot(brent.iloc[-12:,0], label = 'Mensal', linewidth=2, marker='o', color='blue')
#fig.set_title('Evolução do Preço por Barril de Petróleo Bruto Tipo Brent - Em US$ (FOB)')
fig.set_xlabel('Meses', fontsize=15)
fig.set_ylabel('Preço (US$)', fontsize=15)
plt.yticks(fontsize=15)
plt.xticks(fontsize=15)

plt.grid(False)

str_month_list = list(brent.index[-12:].strftime("%m/%Y"))
plt.xticks(brent.iloc[-12:].index,str_month_list, fontsize=15)

# Valores em cada ponto do gráfico
for x,y in zip(brent['Preço - Brent (FOB)'][-12:].index, brent['Preço - Brent (FOB)'][-12:].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                (x,y), # this is the point to label
                textcoords="offset points", # how to position the text
                xytext=(0,20), # distance from text to points (x,y)
                ha='center',
                size=15) # horizontal alignment can be left, right or center

plt.savefig('graficos/GRÁFICO 9 - BRENT.svg', dpi=600)

capitais = [
    'ARACAJU',
    'BELEM',
    'BELO HORIZONTE',
    'BRASILIA',
    'CAMPO GRANDE',
    'CURITIBA',
    'FORTALEZA',
    'GOIANIA',
    'PORTO ALEGRE',
    'RECIFE',
    'RIO BRANCO',
    'RIO DE JANEIRO',
    'SALVADOR',
    'SAO LUIS',
    'SAO PAULO',
    'VITORIA'
]

precos_curitiba = precos_mensal.loc[(precos_mensal['MUNICÍPIO'].isin(capitais)\
    &(precos_mensal['PRODUTO']=='OLEO DIESEL S10')
    &(precos_mensal['MUNICÍPIO']=='CURITIBA'))]

x = pd.DataFrame(precos_curitiba.resample('M', on='MÊS').mean().tail(12))

ax, fig = plt.subplots(figsize=(17,9))
fig.plot(x[['PREÇO MÉDIO REVENDA']], marker='o') #variação de preço curitiba

fig.set_xlabel('Meses', fontsize=15)
fig.set_ylabel('Preço (R$)', fontsize=15)
plt.yticks(fontsize=15)
plt.xticks(fontsize=15)
plt.xticks(x.iloc[-12:].index,str_month_list, fontsize=15)

plt.grid(False)

# Valores em cada ponto do gráfico
for x,y in zip(x['PREÇO MÉDIO REVENDA'].index, x['PREÇO MÉDIO REVENDA'].values):

    label = "{:.2f}".format(y)

    plt.annotate(label, # this is the text
                (x,y), # this is the point to label
                textcoords="offset points", # how to position the text
                xytext=(0,20), # distance from text to points (x,y)
                ha='center',
                size=15) # horizontal alignment can be left, right or center

plt.savefig('graficos/GRÁFICO 10 - DIESEL CURITIBA.svg', dpi=600)

preco_diesel_regioes = precos_mensal.loc[(precos_mensal['PRODUTO']=='OLEO DIESEL S10')\
    &(precos_mensal['MUNICÍPIO'].isin(capitais))]

precos_diesel_capitais = precos_mensal[(precos_mensal['PRODUTO'] == 'OLEO DIESEL S10') & (precos_mensal['MUNICÍPIO'].isin(capitais))]

precos_3_meses_capitais = precos_diesel_capitais.pivot_table(index='MUNICÍPIO', values='PREÇO MÉDIO REVENDA', columns='MÊS').iloc[:,-3:]

precos_3_meses_capitais['MEDIA'] = precos_3_meses_capitais.sum(axis=1) / 3

precos_3_meses_capitais_ordenado = precos_3_meses_capitais.sort_values('MEDIA', ascending=False)

workbook = oxl.load_workbook(filename='diesel_regioes_modelo.xlsx')

sheet = workbook.active

#Preenche a tabela das categorias do ipca de curitiba
for i in range(0, 16):
    sheet[f"A{i+4}"] = precos_3_meses_capitais_ordenado.index[i]
    sheet[f"B{i+4}"] = precos_3_meses_capitais_ordenado.iloc[:,0][i]
    sheet[f"C{i+4}"] = precos_3_meses_capitais_ordenado.iloc[:,1][i]
    sheet[f"D{i+4}"] = precos_3_meses_capitais_ordenado.iloc[:,2][i]
    sheet[f"E{i+4}"] = precos_3_meses_capitais_ordenado.iloc[:,3][i]

workbook.save(filename="tabelas/PREÇO DIESEL REGIOES.xlsx")

precos_semanal_regioesbr[precos_semanal_regioesbr['PRODUTO']=='OLEO DIESEL S10'].groupby('REGIÃO').mean()

precos_semanal_regioesbr.pivot_table(index='REGIÃO',values='PREÇO MÉDIO REVENDA',columns='DATA FINAL')















